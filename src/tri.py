"""tri — Phase 1 : segmentation de la base par mode de chauffage.

CLI :
    python -m src.tri data/base.csv --outdir out

Sorties :
- out/segments/<SEGMENT>.csv   (un fichier par segment, triés surface ↑)
- out/_isoles_qualite.csv      (lignes invalides + raison, jamais droppées)
- out/synthese.xlsx            (comptes par segment, dont froid+)

Règles (cf. SPEC.md) :
- GAZ/FIOUL → AIR_EAU · ÉLEC → AIR_AIR · BOIS → AIR_EAU_A_QUALIFIER ·
  déjà-PAC/inconnu → EXCLU.
- « Tel seul » (sans email) exclu : canal email seul en V1.
- Sous-tag froid+ si dernier contact > 365 j.
"""
from __future__ import annotations

import argparse
import csv
import re
from datetime import date, datetime
from pathlib import Path

from . import config as C
from .logging_setup import get_logger
from .models import Contact, InvalidRow, TriResult

logger = get_logger("datareno.tri")

# Normalisation des en-têtes de colonnes -> clés canoniques internes.
HEADER_MAP = {
    "nom": "nom",
    "name": "nom",
    "email": "email",
    "mail": "email",
    "e-mail": "email",
    "courriel": "email",
    "tel": "tel",
    "telephone": "tel",
    "phone": "tel",
    "mobile": "tel",
    "tel_national": "tel",
    "tel national": "tel",
    "tel_e164": "tel",
    "tel e164": "tel",
    "cp": "cp",
    "code postal": "cp",
    "codepostal": "cp",
    "dept": "dept",
    "departement": "dept",
    "chauffage": "chauffage",
    "mode de chauffage": "chauffage",
    "energie": "chauffage",
    "surface": "surface",
    "surface m2": "surface",
    "m2": "surface",
    "campagne": "campagne",
    "source": "campagne",
    "cpn": "campagne",
    "date": "date",
    "date contact": "date",
    "date_contact": "date",
    "derniere date": "date",
    "submitted at": "date",
    # Intitulés réels de la base acquise (formulaire de collecte).
    "quel est votre mode de chauffage actuel?": "chauffage",
    "quel est votre mode de chauffage actuel": "chauffage",
    "quelle est la surface de votre maison ?": "surface",
    "quelle est la surface de votre maison": "surface",
    "eml": "email",
}

# Colonnes des fichiers de segment en sortie.
OUTPUT_FIELDS = [
    "nom", "email", "tel", "cp", "dept", "chauffage", "surface",
    "campagne", "date_contact", "froid_plus", "exclusion_reason",
]

# Colonnes de la liste « à rappeler » (contacts sans email exploitable mais avec téléphone).
PHONE_FIELDS = [
    "nom", "tel", "cp", "dept", "chauffage", "surface", "campagne", "date_contact", "raison",
]


def normalize_header(name: str) -> str:
    key = C.strip_accents(name or "").lower().strip()
    key = " ".join(key.split())  # compacte les espaces multiples
    return HEADER_MAP.get(key, key)


def parse_surface(raw: str) -> float | None:
    """Surface en m². Gère un nombre (« 120 m2 ») comme une fourchette
    (« 110-160 m2 », « 50 - 110 m2 ») : dans ce cas on retient la borne basse,
    ce qui préserve le tri par priorité « surface ↑ »."""
    if not raw:
        return None
    without_unit = re.sub(r"m[²2]", "", str(raw), flags=re.IGNORECASE)
    # Fourchette « a - b » -> borne basse a.
    range_match = re.match(r"\s*([0-9][0-9\s.,]*?)\s*[-–]\s*[0-9]", without_unit)
    token = range_match.group(1) if range_match else without_unit
    cleaned = re.sub(r"[^0-9,.]", "", token).replace(",", ".")
    if not cleaned or cleaned == ".":
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_date(raw: str) -> date | None:
    if not raw:
        return None
    text = raw.strip()
    try:  # gère l'ISO avec heure (« 2023-09-07T10:28:00 ») et sans.
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass
    for fmt in C.DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def is_froid_plus(date_contact: date | None, today: date) -> bool:
    if date_contact is None:
        return False
    return (today - date_contact).days > C.FROID_PLUS_DAYS


def segment_for(email: str, tel: str, chauffage_raw: str) -> tuple[str, str | None]:
    """Retourne (segment, raison_exclusion|None)."""
    if not email:
        return C.SEGMENT_EXCLU, (C.REASON_TEL_SEUL if tel else C.REASON_SANS_EMAIL)
    norm = C.normalize_chauffage(chauffage_raw)
    if norm in C.CHAUFFAGE_DEJA_PAC:
        return C.SEGMENT_EXCLU, C.REASON_DEJA_PAC
    canonical = C.CHAUFFAGE_SYNONYMS.get(norm)
    if canonical and canonical in C.CHAUFFAGE_TO_SEGMENT:
        return C.CHAUFFAGE_TO_SEGMENT[canonical], None
    return C.SEGMENT_EXCLU, C.REASON_CHAUFFAGE_INCONNU


def classify_row(raw: dict[str, str], line: int, today: date) -> Contact | InvalidRow:
    """Classe une ligne brute en Contact (avec segment) ou InvalidRow (isolée)."""
    stripped = {k: (v.strip() if isinstance(v, str) else "") for k, v in raw.items()}

    if not any(stripped.values()):
        return InvalidRow(line=line, reason=C.REASON_LIGNE_VIDE, raw=raw)

    email = stripped.get("email", "")
    tel = stripped.get("tel", "")

    if email and not re.match(C.EMAIL_REGEX, email):
        return InvalidRow(line=line, reason=C.REASON_EMAIL_INVALIDE, raw=raw)

    cp = stripped.get("cp", "")
    dept = stripped.get("dept", "") or (cp[:2] if len(cp) >= 2 else "")
    surface = parse_surface(stripped.get("surface", ""))
    date_contact = parse_date(stripped.get("date", ""))
    segment, reason = segment_for(email, tel, stripped.get("chauffage", ""))

    return Contact(
        nom=stripped.get("nom") or None,
        email=email or None,
        tel=tel or None,
        cp=cp or None,
        dept=dept or None,
        chauffage=stripped.get("chauffage") or None,
        surface=surface,
        campagne=stripped.get("campagne") or None,
        date_contact=date_contact,
        segment=segment,
        froid_plus=is_froid_plus(date_contact, today),
        exclusion_reason=reason,
    )


def _read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Lit le fichier d'entrée (CSV ou XLSX) et normalise les en-têtes."""
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    return _read_csv(path)


def _read_csv(csv_path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Lit le CSV (sniff du séparateur, fallback utf-8 → latin-1)."""
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with csv_path.open("r", encoding=encoding, newline="") as fh:
                sample = fh.read(4096)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(fh, dialect=dialect)
                raw_fields = reader.fieldnames or []
                rows = [dict(r) for r in reader]
            normalized = [normalize_header(f) for f in raw_fields]
            remapped = [
                {normalize_header(k): v for k, v in row.items()} for row in rows
            ]
            return normalized, remapped
        except UnicodeDecodeError:
            continue
    raise ValueError("Impossible de décoder le CSV (utf-8/latin-1).")


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Lit la 1ʳᵉ feuille d'un classeur Excel (en-tête = 1ʳᵉ ligne)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []

    # Indices des colonnes nommées (on ignore les colonnes sans en-tête).
    cols = [(i, normalize_header(str(h))) for i, h in enumerate(header) if h not in (None, "")]
    normalized = [name for _, name in cols]

    def _cell(value: object) -> str:
        if value is None:
            return ""
        if hasattr(value, "isoformat"):  # date / datetime
            return value.isoformat()
        return str(value)

    rows = [{name: _cell(raw[i]) for i, name in cols} for raw in rows_iter]
    wb.close()
    return normalized, rows


def _sort_key(contact: Contact) -> tuple[bool, float]:
    # Priorité surface ↑ ; surfaces inconnues en fin de liste.
    return (contact.surface is None, contact.surface or 0.0)


def run(csv_path: Path, outdir: Path, today: date | None = None) -> TriResult:
    """Exécute le tri complet et écrit les sorties. Retourne un TriResult (sans PII)."""
    today = today or date.today()
    raw_fields, rows = _read_rows(csv_path)

    buckets: dict[str, list[Contact]] = {s: [] for s in C.ALL_SEGMENTS}
    invalids: list[InvalidRow] = []
    duplicates: list[Contact] = []
    seen_emails: set[str] = set()

    for idx, row in enumerate(rows, start=2):  # ligne 1 = en-tête
        result = classify_row(row, idx, today)
        if isinstance(result, InvalidRow):
            invalids.append(result)
            continue
        # Dédup sur email (1ʳᵉ occurrence conservée ; les suivantes isolées).
        if result.email:
            key = result.email.lower()
            if key in seen_emails:
                duplicates.append(result)
                continue
            seen_emails.add(key)
        buckets[result.segment].append(result)

    segments_dir = outdir / "segments"
    segments_dir.mkdir(parents=True, exist_ok=True)

    counts: dict[str, int] = {}
    froids: dict[str, int] = {}
    for segment, contacts in buckets.items():
        contacts.sort(key=_sort_key)
        counts[segment] = len(contacts)
        froids[segment] = sum(1 for c in contacts if c.froid_plus)
        _write_segment_csv(segments_dir / f"{segment}.csv", contacts)

    # Liste « à rappeler » : pas d'email exploitable mais un téléphone présent.
    # (Export informatif — l'outil n'automatise aucun appel ; voir garde-fous CLAUDE.md.)
    phone_only = _collect_phone_only(buckets[C.SEGMENT_EXCLU], invalids)

    _write_invalids_csv(outdir / "_isoles_qualite.csv", invalids, raw_fields)
    _write_segment_csv(outdir / "_doublons_email.csv", duplicates)
    _write_phone_csv(outdir / "_a_rappeler_telephone.csv", phone_only)
    _write_synthese_xlsx(
        outdir / "synthese.xlsx", counts, froids, len(invalids), len(duplicates), len(phone_only)
    )

    logger.info(
        "tri terminé",
        extra={"context": {
            "total": len(rows),
            "counts": counts,
            "froid_plus": froids,
            "isoles": len(invalids),
            "doublons": len(duplicates),
            "a_rappeler": len(phone_only),
        }},
    )

    return TriResult(
        total_rows=len(rows),
        counts_by_segment=counts,
        froid_plus_by_segment=froids,
        invalid_count=len(invalids),
        duplicate_count=len(duplicates),
        phone_only_count=len(phone_only),
    )


def _collect_phone_only(
    exclus: list[Contact], invalids: list[InvalidRow]
) -> list[dict[str, str]]:
    """Contacts joignables par téléphone mais non emailables (tel seul + email invalide)."""
    out: list[dict[str, str]] = []
    for c in exclus:
        if c.exclusion_reason == C.REASON_TEL_SEUL and c.tel:
            out.append({
                "nom": c.nom or "", "tel": c.tel, "cp": c.cp or "", "dept": c.dept or "",
                "chauffage": c.chauffage or "", "surface": "" if c.surface is None else c.surface,
                "campagne": c.campagne or "",
                "date_contact": c.date_contact.isoformat() if c.date_contact else "",
                "raison": c.exclusion_reason or "",
            })
    for inv in invalids:
        s = {k: (v.strip() if isinstance(v, str) else "") for k, v in inv.raw.items()}
        tel = s.get("tel", "")
        if not tel:
            continue
        cp = s.get("cp", "")
        out.append({
            "nom": s.get("nom", ""), "tel": tel, "cp": cp,
            "dept": s.get("dept", "") or (cp[:2] if len(cp) >= 2 else ""),
            "chauffage": s.get("chauffage", ""), "surface": s.get("surface", ""),
            "campagne": s.get("campagne", ""), "date_contact": s.get("date", ""),
            "raison": inv.reason,
        })
    return out


def _write_phone_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=PHONE_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_segment_csv(path: Path, contacts: list[Contact]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        for c in contacts:
            row = c.model_dump()
            row["date_contact"] = c.date_contact.isoformat() if c.date_contact else ""
            writer.writerow({k: row.get(k, "") for k in OUTPUT_FIELDS})


def _write_invalids_csv(
    path: Path, invalids: list[InvalidRow], raw_fields: list[str]
) -> None:
    fieldnames = ["line", "reason"] + raw_fields
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for inv in invalids:
            record = {"line": inv.line, "reason": inv.reason}
            record.update(inv.raw)
            writer.writerow(record)


def _write_synthese_xlsx(
    path: Path,
    counts: dict[str, int],
    froids: dict[str, int],
    invalid_count: int,
    duplicate_count: int,
    phone_only_count: int,
) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"
    ws.append(["Segment", "Contacts", "dont froid+"])
    for segment in C.ALL_SEGMENTS:
        ws.append([segment, counts.get(segment, 0), froids.get(segment, 0)])
    ws.append(["ISOLES_QUALITE", invalid_count, 0])
    ws.append(["DOUBLONS_EMAIL", duplicate_count, 0])
    ws.append(["A_RAPPELER_TEL", phone_only_count, 0])
    wb.save(path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Tri de la base réno par chauffage.")
    parser.add_argument("source", help="Fichier d'entrée (.csv ou .xlsx).")
    parser.add_argument("--outdir", default="out", help="Dossier de sortie (défaut: out).")
    args = parser.parse_args(argv)

    result = run(Path(args.source), Path(args.outdir))
    print(  # noqa: T201 — sortie CLI volontaire
        f"Tri OK — {result.total_rows} lignes · "
        + " · ".join(f"{s}={result.counts_by_segment.get(s, 0)}" for s in C.ALL_SEGMENTS)
        + f" · isolés={result.invalid_count} · doublons={result.duplicate_count}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
